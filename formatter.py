import sqlite3
import pandas as pd
import os
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from datetime import date
from login_helper import db

db_path = None
excel_path = None

academic_year, result_type, department, semester, result_name, course_code = None, None, None, None, None, None

def format_results(db_path, academic_year, result_type, department, semester, result_name, course_code, college_id):  
    if not os.path.exists(db_path):
        print(f"Error: Database file '{db_path}' not found in the current directory.")
        return

    excel_path = os.path.join("downloads", f'{(db_path.split('.')[0]).split("\\")[1]}.xlsx')

    # 1. CONNECT TO DB AND READ TABLES INTO PANDAS DATAFRAMES
    print("Connecting to the database and reading tables...")
    conn = sqlite3.connect(db_path)
    
    students_df = pd.read_sql_query("SELECT * FROM student", conn)
    subjects_df = pd.read_sql_query("SELECT * FROM subject", conn)
    results_df = pd.read_sql_query("SELECT * FROM results", conn)
    
    conn.close()

    students_df['seat_no'] = pd.to_numeric(students_df['seat_no'], errors='coerce')
    results_df['seat_no'] = pd.to_numeric(results_df['seat_no'], errors='coerce')

    # 2. MERGE THE DATA INTO A SINGLE MASTER DATAFRAME
    print("Merging data from the three tables...")
    merged_df = pd.merge(results_df, students_df, on='seat_no')
    full_data_df = pd.merge(merged_df, subjects_df, on='subject_code')

    # 3. PIVOT THE DATA TO CREATE THE WIDE FORMAT
    print("Pivoting data to create the report card format...")
    pivoted_df = full_data_df.pivot_table(
        index=['seat_no', 'name'],
        columns='subject_code',
        values=['total_marks', 'subject_status'],
        aggfunc='first'
    )

    # 4. CLEAN UP AND REORDER THE PIVOTED TABLE COLUMNS
    pivoted_df.columns = pivoted_df.columns.swaplevel(0, 1)
    pivoted_df.sort_index(axis=1, level=0, inplace=True)
    pivoted_df.rename(columns={'total_marks': 'Total Marks', 'subject_status': 'Grade'}, inplace=True)

    ordered_subject_codes = pivoted_df.columns.get_level_values(0).unique()
    new_column_order = []
    for code in ordered_subject_codes:
        new_column_order.append((code, 'Total Marks'))
        new_column_order.append((code, 'Grade'))

    pivoted_df = pivoted_df[new_column_order]
    
    # REMARKS COLUMN
    if 'remarks' in full_data_df.columns:
        full_data_df['remarks'] = full_data_df['remarks'].fillna('')
        
        remarks_df = full_data_df.groupby(['seat_no', 'name'])['remarks'].apply(
            lambda x: ', '.join(x[x != ''].unique())
        ).reset_index()

        remarks_df.set_index(['seat_no', 'name'], inplace=True)

        pivoted_df[('Remarks', '')] = remarks_df['remarks']
        pivoted_df[('Remarks', '')] = pivoted_df[('Remarks', '')].fillna('')
        
    pivoted_df.index.names = ['PRN', 'Student Name']

    # SGPA COLUMN
    if 'sgpa' in full_data_df.columns:
        sgpa_df = full_data_df.groupby(['seat_no', 'name'])['sgpa'].first().reset_index()

        sgpa_df.set_index(['seat_no', 'name'], inplace=True)

        pivoted_df[('SGPA', '')] = sgpa_df['sgpa']
        
        pivoted_df[('SGPA', '')] = pivoted_df[('SGPA', '')].fillna('-')

    # 5. WRITE DATA AND MANUALLY CREATE CUSTOM HEADERS IN EXCEL
    print(f"Writing formatted data to '{excel_path}'...")
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        pivoted_df.to_excel(
            writer, 
            sheet_name='Analysis Report', 
            startrow=12,
            startcol=1,
            index=True,
            header=False
        )

        worksheet = writer.sheets['Analysis Report']

        # --- Define Styles---
        bold_font = Font(bold=True)
        center_align_wrapped = Alignment(horizontal='center', vertical='center', wrap_text=True)
        student_name_align = Alignment(vertical='center', wrap_text=True)

        # --- Create custom headers ---
        for i, name in enumerate(pivoted_df.index.names):
            style_cell(worksheet, 13, i + 2, name, bold_font, center_align_wrapped)

        index_col_count = len(pivoted_df.index.names)
        sub_headers = pivoted_df.columns.get_level_values(1)

        for col_num, value in enumerate(sub_headers, start=index_col_count + 2):
            style_cell(worksheet, 13, col_num, value, bold_font, center_align_wrapped)

        ordered_subject_codes = pivoted_df.columns.get_level_values(0).unique()
        subject_map = subjects_df.set_index('subject_code').to_dict()
        
        
        current_col = index_col_count + 2
        for code in ordered_subject_codes:
            if code == 'Remarks' or code == 'SGPA':
                start_col = current_col

                # Remarks, SGPA column styling
                worksheet.merge_cells(start_row=10, start_column=start_col, end_row=12, end_column=start_col)
                worksheet.merge_cells(start_row=10, start_column=start_col + 1, end_row=12, end_column=start_col + 1)
                style_cell(worksheet, 10, start_col, "Remarks", bold_font, center_align_wrapped)
                style_cell(worksheet, 10, start_col + 1, "SGPA", bold_font, center_align_wrapped)
            else:
                start_col = current_col
                end_col = current_col + 1

                style_cell(worksheet, 10, start_col, code, bold_font, center_align_wrapped)
                worksheet.merge_cells(start_row=10, start_column=start_col, end_row=10, end_column=end_col)

                name = subject_map['subject_name'].get(code, '')
                style_cell(worksheet, 11, start_col, name, bold_font, center_align_wrapped)
                worksheet.merge_cells(start_row=11, start_column=start_col, end_row=11, end_column=end_col)

                prof = subject_map['professor'].get(code, '')
                style_cell(worksheet, 12, start_col, prof, bold_font, center_align_wrapped)
                worksheet.merge_cells(start_row=12, start_column=start_col, end_row=12, end_column=end_col)

                current_col += 2

        for i in range(10, 14):
            worksheet.cell(row=i, column=2).alignment = center_align_wrapped
            worksheet.cell(row=i, column=3).alignment = center_align_wrapped

        worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=9)
        
        style_cell(worksheet, 10, 2, "Subject Code:", bold_font, center_align_wrapped)
        style_cell(worksheet, 11, 2, "Subject Name:", bold_font, center_align_wrapped)
        style_cell(worksheet, 12, 2, "Subject Teacher:", bold_font, center_align_wrapped)

        style_cell(worksheet, 6, 11, "Form No.:", bold_font, student_name_align)
        worksheet.merge_cells(start_row=6, start_column=11, end_row=6, end_column=13)
        worksheet.merge_cells(start_row=6, start_column=14, end_row=6, end_column=16)
        style_cell(worksheet, 6, 14, "AC-11A", bold_font, student_name_align)

        style_cell(worksheet, 7, 11, "Rev. No. | Issue Date:", bold_font, student_name_align)
        worksheet.merge_cells(start_row=7, start_column=11, end_row=7, end_column=13)
        worksheet.merge_cells(start_row=7, start_column=14, end_row=7, end_column=16)
        style_cell(worksheet, 7, 14, f'00 | {date.today().strftime("%b %d, %Y").upper()}', bold_font, student_name_align)

        style_cell(worksheet, 5, 2, "RESULT ANALYSIS", bold_font, center_align_wrapped)
        worksheet.merge_cells(start_row=5, start_column=2, end_row=5, end_column=9)

        style_cell(worksheet, 6, 2, "Academic Year:", bold_font, student_name_align)
        style_cell(worksheet, 6, 3, academic_year, bold_font, student_name_align)
        style_cell(worksheet, 7, 2, "Result Name:", bold_font, student_name_align)

        style_cell(worksheet, 7, 3, result_type, bold_font, student_name_align)
        style_cell(worksheet, 8, 2, "Department:", bold_font, student_name_align)
        style_cell(worksheet, 8, 3, department, bold_font, student_name_align)

        style_cell(worksheet, 6, 6, "Semester:", bold_font, student_name_align)
        style_cell(worksheet, 6, 8, str(semester), bold_font, student_name_align)
        worksheet.merge_cells(start_row=6, start_column=6, end_row=6, end_column=7)
        worksheet.merge_cells(start_row=6, start_column=8, end_row=6, end_column=9)

        if semester <= 2 and semester > 0:
            result_name = "First year"
        elif semester > 2 and semester <= 4:
            result_name = "Second year"
        elif semester > 4 and semester <= 6:
            result_name = "Third year"
        elif semester > 6 and semester <= 8:
            result_name = "Fourth year"
        else:
            result_name = "NOT VALID"

        style_cell(worksheet, 7, 6, "Class:", bold_font, student_name_align)
        style_cell(worksheet, 7, 8, result_name, bold_font, student_name_align)
        worksheet.merge_cells(start_row=7, start_column=6, end_row=7, end_column=7)
        worksheet.merge_cells(start_row=7, start_column=8, end_row=7, end_column=9)
        
        style_cell(worksheet, 8, 6, "Course Code:", bold_font, student_name_align)
        style_cell(worksheet, 8, 8, course_code, bold_font, student_name_align)
        worksheet.merge_cells(start_row=8, start_column=6, end_row=8, end_column=7)
        worksheet.merge_cells(start_row=8, start_column=8, end_row=8, end_column=9)

        
        

        data_header_font = Font(bold=True, size=16)

        for cell in worksheet[5]:
            cell.font = data_header_font
            cell.alignment = center_align_wrapped

        # Apply wrapping to student names
        for cell in worksheet['C']:
             if cell.row > 13:
                cell.alignment = student_name_align
        
        # Set a fixed, narrower width for all subject columns
        for i in range(4, worksheet.max_column + 1):
            col_letter = get_column_letter(i)
            worksheet.column_dimensions[col_letter].width = 8
            
        # Set number format for the PRN column
        for cell in worksheet['B']:
            if cell.row >= 14:
                cell.number_format = '0'
        
        
        # Add serial numbers
        style_cell(worksheet, 10, 1, "Sr. No.", bold_font, center_align_wrapped)
        worksheet.merge_cells(start_row=10, start_column=1, end_row=13, end_column=1)
        
        serial_number = 1
        for row_num in range(14, worksheet.max_row + 1):
            style_cell(worksheet, row_num, 1, serial_number, bold_font, center_align_wrapped)
            serial_number += 1

        
        # --- SUMMARY ---
        print("Calculating summary statistics using Excel formulas...")

        # --- 1. Define Ranges and Start Row ---
        start_data_row = 14
        end_data_row = worksheet.max_row
        summary_start_row = end_data_row + 2 # Start summary 2 rows below data

        # --- 2. Write the Summary Row Labels ---
        summary_labels = [
            "Total Pass Students:", 
            "Total Fail Students:", 
            "Total No. Appeared Students:"
        ]
        # These will be written in Column C, starting from summary_start_row

        # --- 3. Loop Through Subjects and Write Formulas ---
        num_subjects = ((worksheet.max_column - 4) // 2)  # Exclude Sr, PRN, Name, Remarks
        
        for i in range(num_subjects):
            # Calculate the column letter for the 'Grade' column of the current subject
            grade_col_num = 5 + (i * 2)
            grade_col_letter = get_column_letter(grade_col_num)
            
            # Define the cell range for this subject's grades (e.g., F14:FXX)
            grade_range = f"{grade_col_letter}{start_data_row}:{grade_col_letter}{end_data_row}"

            # --- Write Formulas for Counts ---
            pass_cell = worksheet.cell(row=summary_start_row, column=grade_col_num - 1)
            pass_cell.value = f'=COUNTIF({grade_range}, "PASS")'
            
            fail_cell = worksheet.cell(row=summary_start_row + 1, column=grade_col_num - 1)
            fail_cell.value = f'=COUNTIF({grade_range}, "FAIL")'
            
            appeared_cell = worksheet.cell(row=summary_start_row + 2, column=grade_col_num - 1)
            appeared_cell.value = f'={pass_cell.coordinate}+{fail_cell.coordinate}'

            # --- Write Formulas for Percentages ---
            pass_perc_cell = worksheet.cell(row=summary_start_row, column=grade_col_num)
            pass_perc_cell.value = f'={pass_cell.coordinate}/{appeared_cell.coordinate}'
            pass_perc_cell.number_format = '0.00%' # Format as percentage

            fail_perc_cell = worksheet.cell(row=summary_start_row + 1, column=grade_col_num)
            fail_perc_cell.value = f'={fail_cell.coordinate}/{appeared_cell.coordinate}'
            fail_perc_cell.number_format = '0.00%'

        # --- 4. Calculate and Write Overall Summary (using Python for simplicity) ---
        sr_no_range = f"A{start_data_row}:A{end_data_row}"
        total_students = f'=MAX({sr_no_range})'
        res_range = f"{get_column_letter(worksheet.max_column - 1)}{start_data_row}:{get_column_letter(worksheet.max_column - 1)}{end_data_row}"
        overall_pass_count = f'=COUNTIF({res_range}, "All Clear")'
        overall_fail_count = f'=COUNTIF({res_range}, "<>All Clear")'

        # Determine the columns for the overall summary at the end of the table
        remarks_col = worksheet.max_column -2
        overall_count_col = worksheet.max_column - 1

        style_cell(worksheet, summary_start_row, remarks_col, "All Clear:", bold_font, center_align_wrapped)
        worksheet.cell(row=summary_start_row, column=overall_count_col, value=overall_pass_count)

        style_cell(worksheet, summary_start_row + 1, remarks_col, "Backlog:", bold_font, center_align_wrapped)
        worksheet.cell(row=summary_start_row + 1, column=overall_count_col, value=overall_fail_count)
        
        worksheet.cell(row=summary_start_row + 2, column=overall_count_col, value=total_students)

        for i in range(0, 3):
            worksheet.cell(row=worksheet.max_row - i, column=worksheet.max_column - 1).font = bold_font
        
        # --- COLOURING ---
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Iterate through all rows and cells that have data and apply the border
        for row in worksheet.iter_rows(min_row=10, max_row=worksheet.max_row - 4, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_align_wrapped
                if cell.value == "FAIL":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif cell.value == "AB":
                    cell.fill = PatternFill(start_color="FFFFB3", end_color="FFFFB3", fill_type="solid")

        for row in worksheet.iter_rows(min_row=worksheet.max_row - 2, max_row=worksheet.max_row, min_col=3, max_col=worksheet.max_column - 1):
            for cell in row:
                cell.border = thin_border
                cell.alignment = center_align_wrapped

        for i, label in enumerate(summary_labels):
            style_cell(worksheet, summary_start_row + i, 3, label, bold_font, student_name_align)
        
        for row in worksheet.iter_rows(min_row=14, max_row=worksheet.max_row, min_col=3, max_col=3):
            for cell in row:
                cell.alignment = student_name_align
         
        draw_outline_border(worksheet, min_row=6, max_row=8, min_col=2, max_col=9)
        draw_outline_border(worksheet, min_row=6, max_row=7, min_col=11, max_col=16)
        draw_outline_border(worksheet, min_row=5, max_row=5, min_col=2, max_col=9)

        colleges = []
        with db("users.db") as conn:
            rows = conn.execute("SELECT name, logo_path FROM college WHERE id = ?", (college_id,)).fetchall()
            colleges = [dict(row) for row in rows]

        worksheet.cell(row=1, column=3, value=colleges[0]['name']).font = Font(bold=True, size=20)
        worksheet.cell(row=1, column=3).alignment = center_align_wrapped

        # --- SET COLUMN WIDTHS AND FINAL FORMATTING ---
        worksheet.column_dimensions['A'].width = 8  # Sr. no
        worksheet.column_dimensions['B'].width = 18  # PRN
        worksheet.column_dimensions['C'].width = 35  # Student Name
        
        worksheet.row_dimensions[11].height = 60  # Subject Name row
        worksheet.row_dimensions[12].height = 50  # Subject Teacher row

        worksheet.row_dimensions[1].height = 110  # Header row
        worksheet.row_dimensions[5].height = 35  # Header row
        
        # Set a fixed width for the Remarks column
        remarks_col_letter = get_column_letter(worksheet.max_column - 1)
        worksheet.column_dimensions[remarks_col_letter].width = 25 # Remarks column

        # Set a fixed width for the SGPA column
        sgpa_col_letter = get_column_letter(worksheet.max_column)
        worksheet.column_dimensions[sgpa_col_letter].width = 15 # SGPA column

        # --- IMAGE ---
        try:
            img = Image(os.path.join('static', colleges[0]['logo_path']))
            img.height = 150
            img.width = 200

            worksheet.add_image(img, 'A1')
        except FileNotFoundError:
            print(f"Error: '{colleges[0]['logo_path']}' image not found! Logo will be skipped.")

    print("\n✅ Process complete! The formatted Excel file has been created.")
    

def draw_outline_border(worksheet, min_row, max_row, min_col, max_col):


    """
    Draws an outline border around a specified range of cells in a worksheet.

    Args:
        worksheet: The openpyxl worksheet object.
        min_row (int): The starting row of the range.
        max_row (int): The ending row of the range.
        min_col (int): The starting column of the range.
        max_col (int): The ending column of the range.
    """
    thin_side = Side(style='thin')

    # Define border styles for each edge and corner
    border_top = Border(top=thin_side)
    border_bottom = Border(bottom=thin_side)
    border_left = Border(left=thin_side)
    border_right = Border(right=thin_side)
    border_top_left = Border(top=thin_side, left=thin_side)
    border_top_right = Border(top=thin_side, right=thin_side)
    border_bottom_left = Border(bottom=thin_side, left=thin_side)
    border_bottom_right = Border(bottom=thin_side, right=thin_side)

    # Loop through the cells in the defined range and apply borders
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            is_min_row = cell.row == min_row
            is_max_row = cell.row == max_row
            is_min_col = cell.column == min_col
            is_max_col = cell.column == max_col

            # Apply corner borders
            if is_min_row and is_min_col:
                cell.border = border_top_left
            elif is_min_row and is_max_col:
                cell.border = border_top_right
            elif is_max_row and is_min_col:
                cell.border = border_bottom_left
            elif is_max_row and is_max_col:
                cell.border = border_bottom_right
            # Apply edge borders
            elif is_min_row:
                cell.border = border_top
            elif is_max_row:
                cell.border = border_bottom
            elif is_min_col:
                cell.border = border_left
            elif is_max_col:
                cell.border = border_right

def style_cell(worksheet, row, col, value, font, alignment):
    """Gets a cell, sets its value, and applies font/alignment styles."""
    cell = worksheet.cell(row=row, column=col, value=value)
    cell.font = font
    cell.alignment = alignment
    return cell